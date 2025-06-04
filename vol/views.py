import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from django.shortcuts import render
from django.http import HttpResponse
from .forms import UploadFileForm
import tempfile
import os

from django.views.decorators.csrf import csrf_exempt

def get_red_indices(result, exp, time_columns):
    red_indices = set()
    for idx, row in result.iterrows():
        for t in exp:
            for col in time_columns:
                if col in result.columns and pd.notnull(row[col]) and pd.notnull(t):
                    # Ensure both values are datetime objects
                    row_time = pd.to_datetime(row[col])
                    exp_time = pd.to_datetime(t)
                    if abs((row_time - exp_time).total_seconds()) < 60:
                        red_indices.add(idx)
                        break
    return red_indices

def process_files(file1, file2, file1_label="File1", file2_label="File2"):
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)
    exp_str = [
        '2025-03-04 17:00:01',
        '2025-03-25 15:41:17',
        '2025-03-26 15:56:40',
        '2025-03-27 12:16:48'
    ]
    exp = [pd.to_datetime(t) for t in exp_str]
    df1['Time'] = pd.to_datetime(df1['Time'], errors='coerce')
    df1['End Time'] = pd.to_datetime(df1['End Time'], errors='coerce')
    df2['Time'] = pd.to_datetime(df2['Time'], errors='coerce')
    df2['End Time'] = pd.to_datetime(df2['End Time'], errors='coerce')
    df1.dropna(subset=['Time', 'End Time', 'Type', 'Position', 'Profit'], inplace=True)
    df2.dropna(subset=['Time', 'End Time', 'Type', 'Position', 'Profit'], inplace=True)

    def find_matching_trades(df1, df2, time_col, type_col, pos_col, profit_col, time_diff=10):
        matches = []
        for idx1, row1 in df1.iterrows():
            opposite_type = 'Sell' if row1[type_col].strip().lower() == 'buy' else 'Buy'
            close_trades = df2[
                (df2[type_col].str.strip().str.lower() == opposite_type.lower()) &
                (
                    (abs((row1[time_col] - df2['Time']).dt.total_seconds()) <= time_diff) |
                    (abs((row1[time_col] - df2['End Time']).dt.total_seconds()) <= time_diff) |
                    (abs((row1['End Time'] - df2['Time']).dt.total_seconds()) <= time_diff) |
                    (abs((row1['End Time'] - df2['End Time']).dt.total_seconds()) <= time_diff)
                )
            ]
            for idx2, row2 in close_trades.iterrows():
                matches.append({
                    'File1_Position': row1[pos_col],
                    'File1_Profit': row1[profit_col],
                    'File1_Time': row1['Time'],
                    'File1_EndTime': row1['End Time'],
                    'File1_Type': row1[type_col],
                    'File2_Position': row2[pos_col],
                    'File2_Profit': row2[profit_col],
                    'File2_Time': row2['Time'],
                    'File2_EndTime': row2['End Time'],
                    'File2_Type': row2[type_col],
                })
        return pd.DataFrame(matches)

    result = find_matching_trades(df1, df2, 'Time', 'Type', 'Position', 'Profit', time_diff=10)
    rename_dict = {
        'File1_Position': f'{file1_label} - Position',
        'File1_Profit': f'{file1_label} - Profit',
        'File1_Time': f'{file1_label} - Time',
        'File1_EndTime': f'{file1_label} - End Time',
        'File1_Type': f'{file1_label} - Type',
        'File2_Position': f'{file2_label} - Position',
        'File2_Profit': f'{file2_label} - Profit',
        'File2_Time': f'{file2_label} - Time',
        'File2_EndTime': f'{file2_label} - End Time',
        'File2_Type': f'{file2_label} - Type',
    }
    result.rename(columns=rename_dict, inplace=True)
    result = result.reset_index(drop=True)
    time_columns = [
        f'{file1_label} - Time', f'{file1_label} - End Time',
        f'{file2_label} - Time', f'{file2_label} - End Time'
    ]
    return result, exp, time_columns

def upload_files(request):
    if request.method == 'POST' and 'download' not in request.POST:
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file1 = request.FILES['file1']
            file2 = request.FILES['file2']
            file1_label = file1.name.rsplit('.', 1)[0]
            file2_label = file2.name.rsplit('.', 1)[0]
            result, exp, time_columns = process_files(file1, file2, file1_label, file2_label)
            red_indices = get_red_indices(result, exp, time_columns)
            # ذخیره نتیجه در سشن برای دانلود بعدی
            request.session['result_data'] = result.to_json(date_format='iso', orient='split')
            request.session['exp'] = [str(t) for t in exp]
            request.session['time_columns'] = time_columns
            return render(request, 'result.html', {
                'columns': result.columns,
                'rows': result.values,
                'red_indices': [int(i) for i in red_indices],
            })
    elif request.method == 'POST' and 'download' in request.POST:
        import json
        result = pd.read_json(request.session['result_data'], orient='split')
        exp = [pd.to_datetime(t) for t in request.session['exp']]
        time_columns = request.session['time_columns']
        red_indices = get_red_indices(result, exp, time_columns)
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp_path = tmp.name
        result.to_excel(tmp_path, index=False)
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        for idx in red_indices:
            for cell in ws[idx+2]:
                cell.fill = red_fill
        wb.save(tmp_path)
        wb.close()
        with open(tmp_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=matched_trades.xlsx'
        os.remove(tmp_path)
        return response
    else:
        form = UploadFileForm()
    return render(request, 'upload.html', {'form': form})